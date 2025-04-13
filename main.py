import csv
import json
import os
import time
from datetime import datetime, timedelta, timezone

import requests
from openpyxl import Workbook, load_workbook

CSV_FILE = "nbu_workers.csv"
XLSX_FILE = "nbu_workers.xlsx"
PROGRESS_FILE = "progress.json"
HEADERS = ["ПІБ", "Посада", "Місце роботи", "ID декларації", "Дата подання", "Зв'язок з РФ", "Причина підозри",
           "Посилання"]

API_LIST_URL = "https://public-api.nazk.gov.ua/v2/documents/list"
API_DOC_URL = "https://public-api.nazk.gov.ua/v2/documents/"


def is_related_to_russia(declaration: dict) -> tuple[bool, str]:
    try:
        # Дані декларанта
        step1 = declaration.get("data", {}).get("step_1", {}).get("data", {})
        nui = step1.get("non_ukraine_identity", {})

        # 1. Проживання або громадянство в Росії
        if step1.get("actual_country") == "180":
            return True, "actual_country == 180"
        if step1.get("country") == "180":
            return True, "country == 180"

        # 2. Іноземний документ від РФ
        for identity in nui.values():
            if identity.get("nui_document_country") == "180":
                return True, "nui_document_country == 180"

        # 3. Пошук у всіх текстових полях step_1
        step1_text = str(step1).lower()
        if any(kw in step1_text for kw in ["росія", "russia", "російська", "russian"]):
            return True, "Текст у step_1 містить згадку про РФ"

        # 4. Пошук по джерелах доходів, нерухомості, активах
        for key in ["step_3", "step_4", "step_5", "step_6", "step_7", "step_8", "step_9"]:
            section = declaration.get("data", {}).get(key, {}).get("data", {})
            section_text = str(section).lower()
            if any(kw in section_text for kw in
                   ["росія", "russia", "російська", "russian", '"country":"180"', '"citizenship":"180"']):
                return True, f"Дані з {key} містять згадку про РФ"

        # 5. Перевірка родичів — step_2
        relatives = declaration.get("data", {}).get("step_2", {}).get("data", {}).get("relatives", {})
        for rel in relatives.values():
            rel_text = str(rel).lower()
            if any(kw in rel_text for kw in
                   ["росія", "russia", "російська", "russian", '"country":"180"', '"citizenship":"180"']):
                return True, "Родич(і) пов'язані з РФ (step_2)"

    except Exception as e:
        return False, f"Помилка: {e}"

    return False, ""


def generate_monthly_ranges_iso(years_back=5):
    today = datetime.now(timezone.utc)
    start_date = today.replace(year=today.year - years_back, day=1)
    ranges = []
    current = start_date

    while current < today:
        next_month = (current.replace(day=28) + timedelta(days=4)).replace(day=1)
        start_str = current.strftime("%Y-%m-%d")
        end_str = (next_month - timedelta(seconds=1)).strftime("%Y-%m-%d")
        ranges.append((start_str, end_str))
        current = next_month
    return ranges


def iso_to_unix_range(start_str, end_str):
    start_unix = int(datetime.strptime(start_str, "%Y-%m-%d").replace(hour=0, minute=0, second=0).timestamp())
    end_unix = int(datetime.strptime(end_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59).timestamp())
    return start_unix, end_unix


def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, 'r') as f:
            return json.load(f)
    return {"completed_ranges": []}

def save_progress(progress):
    with open(PROGRESS_FILE, 'w') as f:
        json.dump(progress, f, indent=2)



def load_existing_ids(csv_path):
    ids = set()
    if os.path.exists(csv_path):
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                ids.add(row["ID декларації"])
    return ids


def write_csv_and_xlsx(row):
    # CSV
    write_header = not os.path.exists(CSV_FILE)
    with open(CSV_FILE, 'a', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS)
        if write_header and os.stat(CSV_FILE).st_size == 0:
            writer.writeheader()
        writer.writerow(row)

    # XLSX
    if not os.path.exists(XLSX_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)
    else:
        wb = load_workbook(XLSX_FILE)
        ws = wb.active
    ws.append([row[h] for h in HEADERS])
    wb.save(XLSX_FILE)


def process_range(start_unix, end_unix, existing_ids):
    page = 1
    while True:
        params = {
            "start_date": start_unix,
            "end_date": end_unix,
            "page": page
        }
        resp = requests.get(API_LIST_URL, params=params)
        if resp.status_code != 200:
            print(f"❌ Помилка HTTP {resp.status_code}")
            return
        data = resp.json().get("data", [])
        if not data:
            break

        for item in data:
            try:
                person_data = item["data"]["step_1"]["data"]
                doc_id = item["id"]
                if doc_id in existing_ids:
                    continue
                if person_data.get("workPlaceEdrpou") == "00032106" or person_data.get(
                        "workPlace").lower() == "національний банк україни":
                    full_decl = requests.get(API_DOC_URL + doc_id).json()
                    related, reason = is_related_to_russia(full_decl)

                    full_name = f"{person_data.get('lastname', '')} {person_data.get('firstname', '')} {person_data.get('middlename', '')}".strip()
                    job_title = person_data.get("workPost", "")
                    workplace = person_data.get("workPlace", "")
                    date_object_from_json = datetime.fromisoformat(item.get("date", ""))
                    date = date_object_from_json.strftime("%d-%m-%Y")
                    link = f"https://public.nazk.gov.ua/declaration/{doc_id}"

                    row = {
                        "ПІБ": full_name,
                        "Посада": job_title,
                        "Місце роботи": workplace,
                        "ID декларації": doc_id,
                        "Дата подання": date,
                        "Зв'язок з РФ": "Так" if related else "Ні",
                        "Причина підозри": reason if related else "",
                        "Посилання": link
                    }

                    write_csv_and_xlsx(row)
                    existing_ids.add(doc_id)
                    icon = "⚠️" if related else "✅"
                    print(f"{icon} {full_name} — {date}: {datetime.now().strftime('%H:%M:%S')}")

                time.sleep(0.1)
            except Exception as e:
                print(f"❗ Помилка обробки: {e}")
        if len(data) < 100:
            break
        page += 1


def main():
    progress = load_progress()
    completed_ranges = set(tuple(r) for r in progress.get("completed_ranges", []))
    all_ranges = generate_monthly_ranges_iso(years_back=1)
    existing_ids = load_existing_ids(CSV_FILE)

    for range_iso in all_ranges:
        if range_iso in completed_ranges:
            print(f"📆 Діапазон {range_iso[0]} — {range_iso[1]} вже оброблений. Пропускаємо...")
            continue

        start_unix, end_unix = iso_to_unix_range(*range_iso)
        print(f"\n📆 Обробка: {range_iso[0]} — {range_iso[1]}")
        process_range(start_unix, end_unix, existing_ids)
        completed_ranges.add(range_iso)
        save_progress({"completed_ranges": list(completed_ranges)})

    print("\n✅ Завершено. Усі діапазони оброблено.")


if __name__ == "__main__":
    main()
